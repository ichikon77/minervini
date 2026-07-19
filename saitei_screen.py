# -*- coding: utf-8 -*-
"""
JPX「裁定取引の状況」自動記録 → HTML出力 → GitHub Pages公開

https://www.jpx.co.jp/markets/statistics-equities/program/index.html
から日次の裁定取引エクセル(.xls)をダウンロードし、
「２．裁定取引に係る現物ポジション」の売り/買いポジションを取得。

- 履歴は saitei_history.json にローカル保存（全期間ぶん蓄積）
- saitei.html を生成して git push（他スクリーナーと同じ方式）
- ページには直近約10営業日分が載っているので、未記録の日だけ追記
  （数日実行し忘れても自動で埋まる）
- 旧エクセル版(jpx_saitei_positions.xlsx)が同フォルダにあれば初回に取り込み
"""

import os
import re
import sys
import json
import time
import subprocess
import datetime

import requests
import xlrd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

INDEX_URL = "https://www.jpx.co.jp/markets/statistics-equities/program/index.html"
BASE_URL = "https://www.jpx.co.jp"

HISTORY_JSON = os.path.join(SCRIPT_DIR, "saitei_history.json")
REPORT_HTML = "saitei.html"
LEGACY_XLSX = os.path.join(SCRIPT_DIR, "jpx_saitei_positions.xlsx")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
}

FIELDS = ["売り_当限", "売り_翌限以降", "売り_合計", "売り_前日比",
          "買い_当限", "買い_翌限以降", "買い_合計", "買い_前日比"]


def log(msg):
    print(f"[{datetime.datetime.now():%Y-%m-%d %H:%M:%S}] {msg}", flush=True)


# -----------------------------------------
# JPXからのデータ取得
# -----------------------------------------
def fetch_xls_links():
    """一覧ページから (日付文字列 'YYYY-MM-DD', xlsのURL) のリストを取得（日付昇順）"""
    r = requests.get(INDEX_URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    links = {}
    for m in re.finditer(r'href="([^"]+/(\d{6})\.xls)"', r.text):
        path, yymmdd = m.group(1), m.group(2)
        yy, mm, dd = int(yymmdd[:2]), int(yymmdd[2:4]), int(yymmdd[4:6])
        d = datetime.date(2000 + yy, mm, dd)
        links[d.isoformat()] = BASE_URL + path
    return sorted(links.items())


def to_number(v):
    """セル値を数値に変換（▲は減少=マイナス扱い、'-'や空はNone）"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "－"):
        return None
    neg = "▲" in s
    s = s.replace("▲", "").strip()
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return None


def parse_positions(xls_bytes):
    """xlsから「２．裁定取引に係る現物ポジション」の数値を抽出"""
    wb = xlrd.open_workbook(file_contents=xls_bytes)
    sh = wb.sheet_by_index(0)

    sec_row = None
    for r in range(sh.nrows):
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if isinstance(v, str) and "裁定取引に係る現物ポジション" in v.replace("　", "").replace(" ", ""):
                sec_row = r
                break
        if sec_row is not None:
            break
    if sec_row is None:
        raise ValueError("「２．裁定取引に係る現物ポジション」のセクションが見つかりません")

    header_row = None
    for r in range(sec_row, min(sec_row + 10, sh.nrows)):
        row_text = "".join(str(sh.cell_value(r, c)) for c in range(sh.ncols))
        if "売りポジション" in row_text and "買いポジション" in row_text:
            header_row = r
            break
    if header_row is None:
        raise ValueError("売り/買いポジションのヘッダ行が見つかりません")

    buy_col = None
    for c in range(sh.ncols):
        if "買いポジション" in str(sh.cell_value(header_row, c)):
            buy_col = c

    def numbers_in_row(r):
        out = []
        for c in range(sh.ncols):
            n = to_number(sh.cell_value(r, c))
            if n is not None:
                out.append((c, n))
        return out

    qty_row = diff_row = None
    for r in range(header_row + 1, min(header_row + 8, sh.nrows)):
        row_text = "".join(str(sh.cell_value(r, c)) for c in range(sh.ncols))
        nums = numbers_in_row(r)
        if qty_row is None and "株" in row_text and len(nums) >= 6:
            qty_row = r
            continue
        if qty_row is not None and "前日比" in row_text and len(nums) >= 6:
            diff_row = r
            break
    if qty_row is None:
        raise ValueError("株数の行が見つかりません")

    def split_sell_buy(r):
        nums = numbers_in_row(r)
        sell = [n for c, n in nums if c < buy_col][:3]
        buy = [n for c, n in nums if c >= buy_col][:3]
        if len(sell) < 3 or len(buy) < 3:
            vals = [n for _, n in nums]
            sell, buy = vals[:3], vals[3:6]
        return sell, buy

    (s_now, s_next, s_total), (b_now, b_next, b_total) = split_sell_buy(qty_row)

    if abs((s_now + s_next) - s_total) > 0.5 or abs((b_now + b_next) - b_total) > 0.5:
        raise ValueError(
            f"合計の整合性エラー: 売り {s_now}+{s_next}!={s_total} / 買い {b_now}+{b_next}!={b_total}"
        )

    rec = {
        "売り_当限": s_now, "売り_翌限以降": s_next, "売り_合計": s_total,
        "買い_当限": b_now, "買い_翌限以降": b_next, "買い_合計": b_total,
        "売り_前日比": None, "買い_前日比": None,
    }
    if diff_row is not None:
        (_, _, sd), (_, _, bd) = split_sell_buy(diff_row)
        rec["売り_前日比"] = sd
        rec["買い_前日比"] = bd
    return rec


# -----------------------------------------
# 履歴（JSON）
# -----------------------------------------
def load_history():
    if os.path.exists(HISTORY_JSON):
        with open(HISTORY_JSON, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_history(hist):
    with open(HISTORY_JSON, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=1)


def import_legacy_xlsx(hist):
    """旧エクセル版の記録が同フォルダにあれば取り込む（初回のみ有効）"""
    if not os.path.exists(LEGACY_XLSX):
        return 0
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0
    ws = load_workbook(LEGACY_XLSX).active
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        d = str(row[0]).strip()[:10].replace("/", "-")
        try:
            datetime.date.fromisoformat(d)
        except ValueError:
            continue
        if d in hist:
            continue
        vals = list(row[1:9]) + [None] * 8
        hist[d] = dict(zip(FIELDS, vals[:8]))
        n += 1
    return n


# -----------------------------------------
# HTML出力
# -----------------------------------------
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>裁定取引 現物ポジション - {latest_date}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #0f172a; color: #e2e8f0; padding: 24px;
  }}
  h1 {{ font-size: 1.5rem; margin-bottom: 4px; color: #f8fafc; }}
  .subtitle {{ color: #94a3b8; font-size: 0.9rem; margin-bottom: 20px; }}
  .badge {{
    display: inline-block; background: #1e40af; color: #bfdbfe;
    border-radius: 12px; padding: 2px 12px; font-size: 0.85rem; margin-left: 8px;
  }}
  .nav {{ display: flex; gap: 12px; margin-bottom: 20px; font-size: 0.85rem; flex-wrap: wrap; }}
  .nav a {{
    color: #60a5fa; text-decoration: none; background: #1e293b;
    padding: 5px 14px; border-radius: 6px; border: 1px solid #334155;
  }}
  .nav a:hover {{ background: #334155; }}
  .nav a.active {{ background: #1e40af; border-color: #3b82f6; color: #bfdbfe; }}
  .table-wrap {{
    overflow: auto;
    max-height: calc(100vh - 180px);
    max-width: 560px;
  }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; }}
  thead th {{
    background: #1e293b; color: #94a3b8; padding: 10px 14px;
    text-align: right; font-weight: 600;
    position: sticky; top: 0; white-space: nowrap; z-index: 2;
  }}
  thead th:first-child {{ text-align: left; }}
  td {{
    padding: 8px 14px; border-bottom: 1px solid #1e293b;
    text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums;
  }}
  td:first-child {{ text-align: left; color: #94a3b8; }}
  tr:hover td {{ background: #16213a; }}
  td.sell {{ color: #b48a8a; }}          /* 売り: 落ち着いたローズグレー */
  td.buy {{ color: #4ade80; }}           /* 買い: 通常は緑 */
  td.buy-warn {{ color: #f87171; font-weight: bold; }}  /* 買い100万千株超: 赤太字 */
  .latest-row td {{ background: rgba(30,64,175,0.18); font-weight: bold; }}
  .updated {{ text-align: left; font-size: 0.78rem; color: #475569; margin-top: 12px; }}
  .note {{ font-size: 0.78rem; color: #64748b; margin-top: 16px; line-height: 1.8; }}
</style>
</head>
<body>
  <div style="font-size:0.75rem; color:#94a3b8; margin-bottom:8px; font-weight:600;"><svg width="16" height="18" viewBox="0 0 32 36" style="vertical-align:-4px; margin-right:3px"><polygon points="3,1 13,9 2,15" fill="#262626"/><polygon points="29,1 19,9 30,15" fill="#262626"/><polygon points="5,4 11,9 4.5,12.5" fill="#c98f52"/><polygon points="27,4 21,9 27.5,12.5" fill="#c98f52"/><ellipse cx="6.5" cy="21" rx="3.2" ry="5" fill="#e8d5b7"/><ellipse cx="25.5" cy="21" rx="3.2" ry="5" fill="#e8d5b7"/><circle cx="16" cy="17" r="11" fill="#262626"/><circle cx="10.5" cy="12.5" r="1.7" fill="#c98f52"/><circle cx="21.5" cy="12.5" r="1.7" fill="#c98f52"/><circle cx="11" cy="16" r="1.6" fill="#0a0a0a"/><circle cx="21" cy="16" r="1.6" fill="#0a0a0a"/><circle cx="11.5" cy="15.4" r="0.55" fill="#e2e8f0"/><circle cx="21.5" cy="15.4" r="0.55" fill="#e2e8f0"/><ellipse cx="16" cy="23" rx="6" ry="4.5" fill="#c98f52"/><ellipse cx="16" cy="21" rx="2.1" ry="1.5" fill="#1a1a1a"/><path d="M12.8,25.5 Q12.3,33 16,35 Q19.7,33 19.2,25.5 Z" fill="#f06292"/><path d="M16,27 L16,33" stroke="#d81b60" stroke-width="0.9" fill="none"/></svg>かぶチワワの分析デッキ（<a href="https://x.com/kabuchiwa" style="color:#60a5fa; text-decoration:none;">@kabuchiwa</a>）</div>
  <nav class="nav">
    <a href="minervini_report_v2.html">米国株 (Minervini)</a>
    <a href="haitou.html">日本株 (配当)</a>
    <a href="jpminervini.html">日本株 (Minervini)</a>
    <a href="saitei.html" class="active">裁定取引</a>
    <a href="totan.html">日銀利上げ確率</a>
    <a href="fedwatch.html">FRB利上げ確率</a>
    <a href="kinri.html">金利と為替</a>
    <a href="daikin.html">売買代金</a>
    <a href="shinyou.html">信用評価率</a>
    <a href="shutai.html">投資主体別</a>
    <a href="gaikoku.html">海外投資家</a>
    <a href="touraku.html">騰落レシオ</a>
    <a href="karauri.html">空売り比率</a>
    <a href="riron.html">日経理論株価</a>
    <a href="spriron.html">SP500理論株価</a>
    <a href="flow.html">資金フロー</a>
    <a href="map.html">デッキの見方</a>
  </nav>
  <h1>裁定取引に係る現物ポジション</h1>
  <p class="subtitle">最終更新: {updated} | 出所: 東京証券取引所「裁定取引の状況」 | 単位: 千株</p>
  <div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th>日付</th>
        <th>売り合計</th>
        <th>買い合計</th>
      </tr>
    </thead>
    <tbody>
{rows}
    </tbody>
  </table>
  </div>
  <p class="note">
    ・「２．裁定取引に係る現物ポジション」の売り/買いポジション株数の合計（当限＋翌限以降）。<br>
    ・JPXは毎営業日16時頃に前々営業日分を公表。<br>
    ・<a href="{jpx_url}" style="color:#60a5fa">JPX 裁定取引ページ</a><br>
    ・<a href="https://gendai.media/articles/-/158923?imp=0" style="color:#60a5fa">清原氏インタビュー記事</a>
  </p>
  <p class="updated">最終更新: {updated}</p>
</body>
</html>
"""


def fmt_num(v):
    if v is None:
        return "-"
    return f"{v:,.0f}"


def generate_html(hist):
    dates = sorted(hist.keys(), reverse=True)  # 最新が上
    rows = []
    for i, d in enumerate(dates):
        rec = hist[d]
        cls = ' class="latest-row"' if i == 0 else ""
        dd = d.replace("-", "/")
        buy = rec.get("買い_合計")
        buy_cls = "buy-warn" if (buy is not None and buy > 1_000_000) else "buy"
        rows.append(
            f'      <tr{cls}><td>{dd}</td>'
            f'<td class="sell">{fmt_num(rec.get("売り_合計"))}</td>'
            f'<td class="{buy_cls}">{fmt_num(buy)}</td></tr>'
        )
    html = HTML_TEMPLATE.format(
        latest_date=dates[0].replace("-", "/") if dates else "-",
        count=len(dates),
        rows="\n".join(rows),
        jpx_url=INDEX_URL,
        updated=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    path = os.path.join(SCRIPT_DIR, REPORT_HTML)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"HTML出力: {path}")


# -----------------------------------------
# GitHub Pages 自動 push（他スクリーナーと同方式）
# -----------------------------------------
def push_to_github(report_filename):
    log("GitHub Pages に公開中...")
    today = datetime.date.today().isoformat()
    subprocess.run(["git", "-C", SCRIPT_DIR, "add", report_filename,
                    os.path.basename(HISTORY_JSON), ".gitignore",
                    "saitei_screen.py", "saitei_run.bat"], check=True)
    result = subprocess.run(
        ["git", "-C", SCRIPT_DIR, "commit", "-m", "update saitei report " + today],
        capture_output=True,
    )
    if result.returncode != 0:
        msg = result.stdout.decode(errors="ignore") + result.stderr.decode(errors="ignore")
        if "nothing to commit" in msg:
            log("  commit skip (already committed)")
        else:
            log("  commit failed: " + msg)
            return

    for attempt in range(1, 6):
        try:
            subprocess.run(["git", "-C", SCRIPT_DIR, "pull", "--rebase", "--autostash"], check=True)
            subprocess.run(["git", "-C", SCRIPT_DIR, "push"],            check=True)
            log("  Done: https://ichikon77.github.io/minervini/saitei.html")
            return
        except subprocess.CalledProcessError as e:
            log(f"  push failed (attempt {attempt}/5): {e}")
            time.sleep(10)
    log("  push failed finally")


# -----------------------------------------
# main
# -----------------------------------------
def main():
    log("JPX 裁定取引の状況 チェック開始")

    hist = load_history()

    imported = import_legacy_xlsx(hist)
    if imported:
        log(f"旧エクセル版から {imported} 日分を取り込みました")

    try:
        links = fetch_xls_links()
    except Exception as e:
        log(f"エラー: 一覧ページの取得に失敗しました: {e}")
        return

    if not links:
        log("エラー: xlsリンクが見つかりませんでした（ページ構造が変わった可能性）")
        return

    log(f"ページ上のデータ: {len(links)}日分 ({links[0][0]} ～ {links[-1][0]})")

    added = 0
    for d, url in links:
        if d in hist:
            continue
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            hist[d] = parse_positions(r.content)
        except Exception as e:
            log(f"エラー: {d} の取得/解析に失敗: {e}")
            continue
        added += 1
        log(f"追記: {d}  売り合計={hist[d]['売り_合計']:,.0f} / 買い合計={hist[d]['買い_合計']:,.0f} 千株")

    if added or imported:
        save_history(hist)
        log(f"履歴保存: {HISTORY_JSON}（計 {len(hist)} 日分）")

    if not hist:
        log("データがないためHTMLは生成しません")
        return

    generate_html(hist)

    if "--nopush" in sys.argv:
        log("--nopush 指定のため git push はスキップ")
    else:
        push_to_github(REPORT_HTML)

    log("完了")


if __name__ == "__main__":
    main()
